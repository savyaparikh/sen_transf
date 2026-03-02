"""
Intraday Dealer Sentiment Tracker — Real-Time Ready
=====================================================
Two-step architecture:
  1. pre_market(df_history) — run once before market open, computes prior close
  2. on_tick(tick)          — call on every incoming tick, returns bps delta

Output: call get_chart_data() or to_excel() anytime to get current state.

Field names expected (same as your feed):
  SECURITY_ID, Timestamp, OFFER_QUANTITY, PRICE_VALUE,
  ALT_PRICE_VALUE, MINE_FLAG, SETTLE_DATE, ECN, ECN_MSG_ID
"""

import pandas as pd
import numpy as np
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')


# =============================================================================
# QTY BUCKETS
# =============================================================================
QTY_BUCKETS = [
    (0,        50_000,    '<50K'),
    (50_000,   100_000,   '50-100K'),
    (100_000,  150_000,   '100-150K'),
    (150_000,  200_000,   '150-200K'),
    (200_000,  500_000,   '200-500K'),
    (500_000,  1_000_000, '500K-1MM'),
    (1_000_000, float('inf'), '1MM+'),
]

QTY_BUCKET_ORDER = {label: i for i, (_, _, label) in enumerate(QTY_BUCKETS)}


def assign_qty_bucket(qty):
    for low, high, label in QTY_BUCKETS:
        if low <= qty < high:
            return label
    return '1MM+'


# =============================================================================
# MAIN CLASS
# =============================================================================
class IntraDaySentiment:

    def __init__(self, bucket_minutes=5, outlier_iqr_multiplier=2.5):
        self.bucket_minutes = bucket_minutes
        self.outlier_iqr_mult = outlier_iqr_multiplier

        # Set by pre_market()
        self.prior_close = {}       # key: (SECURITY_ID, qty_bucket) -> avg_close
        self.prior_close_df = None  # full DataFrame for reference

        # Built up by on_tick()
        self.ticks = []             # list of processed tick dicts
        self._bucket_cache = {}     # key: (qty_bucket, time_bucket) -> list of bps deltas

    # =================================================================
    # STEP 1: PRE-MARKET — run once with prior day's data
    # =================================================================
    def pre_market(self, df_history):
        """
        Compute prior day's close from historical data.
        Pass in at minimum the last trading day's data.
        If you pass multiple days, it uses the most recent day.

        Parameters
        ----------
        df_history : pd.DataFrame
            Historical tick data with standard field names.
        """
        df = df_history.copy()
        df['Date'] = df['Timestamp'].dt.date
        df['ECN_MSG_ID'] = df['ECN_MSG_ID'].astype(str)
        df['qty_bucket'] = df['OFFER_QUANTITY'].apply(assign_qty_bucket)
        df = df.sort_values('Timestamp')

        # Use most recent date as "prior day"
        prior_date = df['Date'].max()
        df_prior = df[df['Date'] == prior_date]

        # Remove outliers
        df_prior = self._remove_outliers(df_prior)

        # Each dealer's last offer per (CUSIP, qty_bucket)
        last_per_dealer = (
            df_prior.sort_values('Timestamp')
            .groupby(['SECURITY_ID', 'qty_bucket', 'ECN_MSG_ID'])
            .agg(last_price=('ALT_PRICE_VALUE', 'last'))
            .reset_index()
        )

        # Average across dealers per (CUSIP, qty_bucket)
        closes = (
            last_per_dealer
            .groupby(['SECURITY_ID', 'qty_bucket'])
            .agg(avg_close=('last_price', 'mean'), num_dealers=('ECN_MSG_ID', 'nunique'))
            .reset_index()
        )

        # Store as lookup dict
        self.prior_close = {}
        for _, row in closes.iterrows():
            key = (row['SECURITY_ID'], row['qty_bucket'])
            self.prior_close[key] = row['avg_close']

        self.prior_close_df = closes
        self.prior_close_df['Date'] = prior_date

        # Reset intraday state
        self.ticks = []
        self._bucket_cache = {}

        print(f"  Pre-market done: {prior_date}")
        print(f"  {len(self.prior_close)} (CUSIP, qty_bucket) baselines loaded:")
        for (sec, qty), close in sorted(self.prior_close.items()):
            print(f"    {sec} / {qty:>10} -> {close:.3f}")

    # =================================================================
    # STEP 2: ON_TICK — call on every incoming tick
    # =================================================================
    def on_tick(self, tick):
        """
        Process a single incoming tick.

        Parameters
        ----------
        tick : dict or pd.Series
            Must have: SECURITY_ID, Timestamp, OFFER_QUANTITY,
            ALT_PRICE_VALUE, ECN_MSG_ID, ECN

        Returns
        -------
        dict with: qty_bucket, time_bucket, bps_delta, or None if no baseline
        """
        if isinstance(tick, pd.Series):
            tick = tick.to_dict()

        sec_id = tick['SECURITY_ID']
        qty = tick['OFFER_QUANTITY']
        price = tick['ALT_PRICE_VALUE']
        ts = tick['Timestamp']
        if isinstance(ts, str):
            ts = pd.Timestamp(ts)

        qty_bucket = assign_qty_bucket(qty)

        # Look up this CUSIP's prior close
        key = (sec_id, qty_bucket)
        if key not in self.prior_close:
            return None

        prior = self.prior_close[key]

        # Compute bps delta (normalized to this CUSIP)
        bps_delta = round((price - prior) * 100, 2)

        # Time bucket
        time_bucket = ts.floor(f'{self.bucket_minutes}min').strftime('%H:%M')

        # Store tick
        processed = {
            'SECURITY_ID': sec_id,
            'qty_bucket': qty_bucket,
            'time_bucket': time_bucket,
            'Timestamp': ts,
            'ALT_PRICE_VALUE': price,
            'prior_close': prior,
            'bps_delta': bps_delta,
            'ECN_MSG_ID': str(tick.get('ECN_MSG_ID', '')),
            'ECN': tick.get('ECN', ''),
        }
        self.ticks.append(processed)

        # Update bucket cache
        cache_key = (qty_bucket, time_bucket)
        if cache_key not in self._bucket_cache:
            self._bucket_cache[cache_key] = []
        self._bucket_cache[cache_key].append(processed)

        return processed

    def on_tick_batch(self, df):
        """
        Process a batch of ticks (DataFrame).
        Convenience wrapper around on_tick().
        """
        results = []
        for _, row in df.iterrows():
            result = self.on_tick(row)
            if result is not None:
                results.append(result)
        return results

    # =================================================================
    # GET AGGREGATED CHART DATA (call anytime)
    # =================================================================
    def get_chart_data(self):
        """
        Aggregate all ticks received so far into chart-ready data.
        Dealer-weighted, CUSIP-normalized.

        Returns
        -------
        pd.DataFrame: qty_bucket, time_bucket, avg_dod_bps, min_dod_bps,
                       max_dod_bps, tick_count, num_cusips, num_dealers
        """
        if not self.ticks:
            return pd.DataFrame()

        df = pd.DataFrame(self.ticks)

        # Per CUSIP per time bucket
        cusip_level = (
            df.groupby(['qty_bucket', 'time_bucket', 'SECURITY_ID'])
            .agg(
                cusip_avg_bps=('bps_delta', 'mean'),
                cusip_min_bps=('bps_delta', 'min'),
                cusip_max_bps=('bps_delta', 'max'),
                cusip_ticks=('bps_delta', 'count'),
                cusip_dealers=('ECN_MSG_ID', 'nunique'),
            )
            .reset_index()
        )

        # Dealer-weighted aggregation across CUSIPs
        def weighted_agg(grp):
            w = grp['cusip_dealers']
            total_w = w.sum()
            if total_w == 0:
                w = np.ones(len(grp)) / len(grp)
            else:
                w = w / total_w
            return pd.Series({
                'avg_dod_bps': round((grp['cusip_avg_bps'] * w).sum(), 2),
                'min_dod_bps': round(grp['cusip_min_bps'].min(), 2),
                'max_dod_bps': round(grp['cusip_max_bps'].max(), 2),
                'tick_count': int(grp['cusip_ticks'].sum()),
                'num_cusips': len(grp),
                'num_dealers': int(grp['cusip_dealers'].sum()),
            })

        result = (
            cusip_level.groupby(['qty_bucket', 'time_bucket'])
            .apply(weighted_agg).reset_index()
        )
        result['_qty_sort'] = result['qty_bucket'].map(QTY_BUCKET_ORDER)
        return result.sort_values(['_qty_sort', 'time_bucket']).reset_index(drop=True)

    def get_latest_reading(self):
        """
        Quick snapshot: latest bps reading per qty bucket.

        Returns
        -------
        dict: qty_bucket -> {time_bucket, avg_dod_bps, tick_count}
        """
        cd = self.get_chart_data()
        if cd.empty:
            return {}

        latest = {}
        for qty, grp in cd.groupby('qty_bucket'):
            last = grp.sort_values('time_bucket').iloc[-1]
            latest[qty] = {
                'time_bucket': last['time_bucket'],
                'avg_dod_bps': last['avg_dod_bps'],
                'tick_count': int(last['tick_count']),
                'num_dealers': int(last['num_dealers']),
            }
        return latest

    def get_raw_ticks(self):
        """Return all processed ticks as DataFrame."""
        return pd.DataFrame(self.ticks) if self.ticks else pd.DataFrame()

    def get_prior_closes(self):
        """Return prior close baselines."""
        return self.prior_close_df

    def reset_intraday(self):
        """Clear today's ticks. Call at end of day before next pre_market()."""
        self.ticks = []
        self._bucket_cache = {}

    # =================================================================
    # OUTLIER REMOVAL (internal)
    # =================================================================
    def _remove_outliers(self, df):
        clean_mask = pd.Series(True, index=df.index)
        for (sec, qty), grp in df.groupby(['SECURITY_ID', 'qty_bucket']):
            if len(grp) < 4:
                continue
            prices = grp['ALT_PRICE_VALUE']
            q1, q3 = prices.quantile(0.25), prices.quantile(0.75)
            iqr = q3 - q1
            if iqr == 0:
                continue
            lower, upper = q1 - self.outlier_iqr_mult * iqr, q3 + self.outlier_iqr_mult * iqr
            clean_mask.loc[grp[(prices < lower) | (prices > upper)].index] = False
        return df[clean_mask].reset_index(drop=True)

    # =================================================================
    # EXCEL OUTPUT
    # =================================================================
    def to_excel(self, filepath):
        cd = self.get_chart_data()
        if cd.empty:
            print("  No ticks to export.")
            return

        wb = Workbook()
        hdr_font = Font(bold=True, size=11, color='FFFFFF')
        hdr_fill = PatternFill('solid', fgColor='2F5496')
        thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

        # --- Sheet 1: Sentiment Data ---
        ws = wb.active
        ws.title = "Sentiment Data"
        headers = ['Qty Bucket', 'Time Bucket', 'Avg DoD (bps)', 'Min (bps)',
                   'Max (bps)', 'Ticks', 'CUSIPs', 'Dealers']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font, c.fill, c.border = hdr_font, hdr_fill, thin
            c.alignment = Alignment(horizontal='center')

        for ri, (_, r) in enumerate(cd.iterrows(), 2):
            vals = [r['qty_bucket'], r['time_bucket'], r['avg_dod_bps'],
                    r['min_dod_bps'], r['max_dod_bps'],
                    r['tick_count'], r['num_cusips'], r['num_dealers']]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=ri, column=col, value=val)
                c.border = thin
                if col in [3, 4, 5]:
                    c.number_format = '+0.00;-0.00;0.00'
                c.alignment = Alignment(horizontal='center')

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 16

        # --- Sheet 2: Chart ---
        cs = wb.create_sheet("Chart")
        qty_buckets = sorted(cd['qty_bucket'].unique(), key=lambda x: QTY_BUCKET_ORDER.get(x, 99))
        all_times = sorted(cd['time_bucket'].unique())

        cs.cell(row=1, column=1, value='Time').font = hdr_font
        cs.cell(row=1, column=1).fill = hdr_fill
        cs.cell(row=1, column=1).border = thin
        for qi, qty in enumerate(qty_buckets, 2):
            c = cs.cell(row=1, column=qi, value=qty)
            c.font, c.fill, c.border = hdr_font, hdr_fill, thin
            c.alignment = Alignment(horizontal='center')

        for ri, t in enumerate(all_times, 2):
            cs.cell(row=ri, column=1, value=t).border = thin
            for qi, qty in enumerate(qty_buckets, 2):
                match = cd[(cd['time_bucket'] == t) & (cd['qty_bucket'] == qty)]
                val = match['avg_dod_bps'].iloc[0] if not match.empty else None
                c = cs.cell(row=ri, column=qi, value=val)
                c.border = thin
                if val is not None:
                    c.number_format = '+0.00;-0.00;0.00'

        cs.column_dimensions['A'].width = 10
        for qi in range(2, len(qty_buckets) + 2):
            cs.column_dimensions[get_column_letter(qi)].width = 14

        chart = LineChart()
        chart.title = "Dealer Sentiment — bps Δ from Prior Close"
        chart.x_axis.title = "Timestamp"
        chart.y_axis.title = "bps"
        chart.style = 10
        chart.width = 28
        chart.height = 14

        n_rows = len(all_times)
        cats = Reference(cs, min_col=1, min_row=2, max_row=1 + n_rows)
        chart.set_categories(cats)

        colors_hex = {
            '<50K': '94A3B8', '50-100K': '06B6D4', '100-150K': '8B5CF6',
            '150-200K': 'F97316', '200-500K': '10B981', '500K-1MM': 'EC4899', '1MM+': 'F59E0B',
        }
        for qi, qty in enumerate(qty_buckets, 2):
            vals = Reference(cs, min_col=qi, min_row=1, max_row=1 + n_rows)
            chart.add_data(vals, titles_from_data=True)
            chart.series[-1].graphicalProperties.line.width = 22000
            chart.series[-1].graphicalProperties.line.solidFill = colors_hex.get(qty, '64748B')

        cs.add_chart(chart, f"A{n_rows + 4}")

        # --- Sheet 3: Prior Closes ---
        ws3 = wb.create_sheet("Prior Closes")
        ch = ['SECURITY_ID', 'Qty Bucket', 'Avg Close', 'Num Dealers']
        for col, h in enumerate(ch, 1):
            c = ws3.cell(row=1, column=col, value=h)
            c.font, c.fill, c.border = hdr_font, hdr_fill, thin
            c.alignment = Alignment(horizontal='center')
        if self.prior_close_df is not None:
            for ri, (_, r) in enumerate(self.prior_close_df.iterrows(), 2):
                vals = [r['SECURITY_ID'], r['qty_bucket'], r['avg_close'], r['num_dealers']]
                for col, val in enumerate(vals, 1):
                    c = ws3.cell(row=ri, column=col, value=val)
                    c.border = thin
                    if col == 3: c.number_format = '0.000'
                    c.alignment = Alignment(horizontal='center')
        for col in range(1, 5):
            ws3.column_dimensions[get_column_letter(col)].width = 16

        wb.save(filepath)
        print(f"  Saved: {filepath}")

    # =================================================================
    # SUMMARY
    # =================================================================
    def summary(self):
        cd = self.get_chart_data()
        if cd.empty:
            print("No ticks processed yet.")
            return
        print("=" * 65)
        print("INTRADAY SENTIMENT")
        print("=" * 65)
        avail = sorted(cd['qty_bucket'].unique(), key=lambda x: QTY_BUCKET_ORDER.get(x, 99))
        print(f"  Qty buckets: {avail} | Total ticks: {len(self.ticks)}")
        print(f"\n  {'Qty':<12} {'Open':>7} {'Latest':>7} {'Low':>7} {'High':>7} {'Ticks':>6}")
        print(f"  {'─'*12} {'─'*7} {'─'*7} {'─'*7} {'─'*7} {'─'*6}")
        for _, grp in cd.groupby('_qty_sort'):
            g = grp.sort_values('time_bucket')
            print(f"  {g.iloc[0]['qty_bucket']:<12} {g.iloc[0]['avg_dod_bps']:>+7.2f} "
                  f"{g.iloc[-1]['avg_dod_bps']:>+7.2f} {g['min_dod_bps'].min():>+7.2f} "
                  f"{g['max_dod_bps'].max():>+7.2f} {int(g['tick_count'].sum()):>6}")


# =============================================================================
# EXAMPLE USAGE
# =============================================================================
if __name__ == '__main__':

    INPUT_PATH = r"C:\Users\savya\OneDrive\Desktop\loading\test_dealer_extended_3secids.xlsx"
    OUTPUT_PATH = r"C:\Users\savya\OneDrive\Desktop\loading\dealer_sentiment_output.xlsx"

    df = pd.read_excel(INPUT_PATH)
    df['ECN_MSG_ID'] = df['ECN_MSG_ID'].astype(str)
    df['Date'] = df['Timestamp'].dt.date

    dates = sorted(df['Date'].unique())
    prior_day = dates[-2]   # second-to-last day as "yesterday"
    today = dates[-1]       # last day as "today"

    print(f"Simulating: prior day = {prior_day}, today = {today}\n")

    # -------------------------------------------
    # 1. PRE-MARKET: compute yesterday's close
    # -------------------------------------------
    model = IntraDaySentiment(bucket_minutes=5)
    model.pre_market(df[df['Date'] == prior_day])

    # -------------------------------------------
    # 2. INTRADAY: feed today's ticks one by one
    # -------------------------------------------
    todays_ticks = df[df['Date'] == today].sort_values('Timestamp')

    print(f"\nProcessing {len(todays_ticks)} ticks...\n")
    for _, tick in todays_ticks.iterrows():
        result = model.on_tick(tick)
        if result:
            print(f"  {result['time_bucket']}  {result['SECURITY_ID']}  "
                  f"{result['qty_bucket']:>10}  {result['bps_delta']:>+6.2f} bps")

    # -------------------------------------------
    # 3. GET RESULTS
    # -------------------------------------------
    print()
    model.summary()

    print("\nLatest reading per qty bucket:")
    for qty, reading in model.get_latest_reading().items():
        print(f"  {qty:>10}: {reading['avg_dod_bps']:>+.2f} bps at {reading['time_bucket']}")

    print("\nWriting Excel...")
    model.to_excel(OUTPUT_PATH)

    # -------------------------------------------
    # END OF DAY: reset for next day
    # -------------------------------------------
    # model.reset_intraday()
    # model.pre_market(df[df['Date'] == today])  # today becomes prior day
