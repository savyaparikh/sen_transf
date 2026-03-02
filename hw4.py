"""
Intraday Dealer Sentiment — Dealer-Linked (v6)
================================================
Uses Hungarian algorithm to link today's ECN_MSG_IDs to yesterday's,
then tracks each dealer's own bps change from their own prior offer.

Architecture:
  1. pre_market(df_history)  — builds yesterday's dealer sessions, runs outlier filter
  2. on_tick(tick)           — processes tick, links new ECN_MSG_IDs via Hungarian
  3. get_chart_data()        — aggregates linked dealers' bps into time buckets

Output: time_bucket, qty_bucket, avg_dod_bps (same as before, ready to plot)
"""

import pandas as pd
import numpy as np
from scipy.optimize import linear_sum_assignment
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')


# =============================================================================
# QTY BUCKETS
# =============================================================================
QTY_BUCKETS = [
    (0,        50_000,    '<50K'),
    (50_000,   100_000,   '50-100K'),
    (100_000,  150_000,   '100-150K'),
    (150_000,  200_000,   '150-200K'),
    (200_000,  500_000,   '200-500K'),
    (500_000,  1_000_000, '500K-1MM'),
    (1_000_000, float('inf'), '1MM+'),
]

QTY_BUCKET_ORDER = {label: i for i, (_, _, label) in enumerate(QTY_BUCKETS)}


def assign_qty_bucket(qty):
    for low, high, label in QTY_BUCKETS:
        if low <= qty < high:
            return label
    return '1MM+'


# =============================================================================
# MAIN CLASS
# =============================================================================
class IntraDaySentiment:

    def __init__(self, bucket_minutes=5, outlier_iqr_multiplier=2.5,
                 match_cost_threshold=50.0):
        self.bucket_minutes = bucket_minutes
        self.outlier_iqr_mult = outlier_iqr_multiplier
        self.match_cost_threshold = match_cost_threshold

        # Set by pre_market()
        self.prior_sessions = []        # list of dicts: yesterday's dealer sessions
        self.prior_sessions_df = None

        # Built by on_tick()
        self.today_sessions = {}        # ECN_MSG_ID -> running session state
        self.ticks = []                 # all processed ticks
        self.dealer_links = {}          # today ECN_MSG_ID -> yesterday ECN_MSG_ID
        self._matched_groups = {}       # cache: (CUSIP, qty_bucket) -> already matched flag

    # =================================================================
    # STEP 1: PRE-MARKET
    # =================================================================
    def pre_market(self, df_history):
        """
        Build yesterday's dealer sessions from historical data.
        Pass in the prior day's data (or multiple days — uses most recent).
        """
        df = df_history.copy()
        df['Date'] = df['Timestamp'].dt.date
        df['ECN_MSG_ID'] = df['ECN_MSG_ID'].astype(str)
        df['qty_bucket'] = df['OFFER_QUANTITY'].apply(assign_qty_bucket)
        df = df.sort_values('Timestamp')

        prior_date = df['Date'].max()
        df_prior = df[df['Date'] == prior_date]
        df_prior = self._remove_outliers(df_prior)

        # Build one session per ECN_MSG_ID per (CUSIP, qty_bucket)
        sessions = []
        for (sec, qty, msg_id), grp in df_prior.groupby(['SECURITY_ID', 'qty_bucket', 'ECN_MSG_ID']):
            g = grp.sort_values('Timestamp')
            sessions.append({
                'SECURITY_ID': sec,
                'qty_bucket': qty,
                'ECN_MSG_ID': msg_id,
                'ECN': g['ECN'].iloc[0],
                'first_price': g['ALT_PRICE_VALUE'].iloc[0],
                'last_price': g['ALT_PRICE_VALUE'].iloc[-1],
                'first_post_hour': g['Timestamp'].iloc[0].hour + g['Timestamp'].iloc[0].minute / 60,
                'num_updates': len(g),
            })

        self.prior_sessions = sessions
        self.prior_sessions_df = pd.DataFrame(sessions)

        # Compute price rank within each (CUSIP, qty_bucket)
        if not self.prior_sessions_df.empty:
            self.prior_sessions_df['price_rank'] = (
                self.prior_sessions_df.groupby(['SECURITY_ID', 'qty_bucket'])['last_price']
                .rank(method='dense').astype(int)
            )
            # Update dicts
            for i, row in self.prior_sessions_df.iterrows():
                self.prior_sessions[i]['price_rank'] = row['price_rank']

        # Reset intraday state
        self.today_sessions = {}
        self.ticks = []
        self.dealer_links = {}
        self._matched_groups = {}

        print(f"  Pre-market done: {prior_date}")
        print(f"  {len(sessions)} dealer sessions loaded across "
              f"{df_prior['SECURITY_ID'].nunique()} CUSIPs")

    # =================================================================
    # STEP 2: ON_TICK
    # =================================================================
    def on_tick(self, tick):
        """
        Process one incoming tick.
        If this is a new ECN_MSG_ID, triggers Hungarian matching
        for that (CUSIP, qty_bucket) group.

        Returns dict with bps_delta, or None if no match found.
        """
        if isinstance(tick, pd.Series):
            tick = tick.to_dict()

        sec_id = tick['SECURITY_ID']
        qty = tick['OFFER_QUANTITY']
        price = tick['ALT_PRICE_VALUE']
        ts = tick['Timestamp']
        ecn = tick.get('ECN', '')
        msg_id = str(tick['ECN_MSG_ID'])

        if isinstance(ts, str):
            ts = pd.Timestamp(ts)

        qty_bucket = assign_qty_bucket(qty)
        time_bucket = ts.floor(f'{self.bucket_minutes}min').strftime('%H:%M')

        # Update today's session for this ECN_MSG_ID
        is_new_session = msg_id not in self.today_sessions
        if is_new_session:
            self.today_sessions[msg_id] = {
                'SECURITY_ID': sec_id,
                'qty_bucket': qty_bucket,
                'ECN_MSG_ID': msg_id,
                'ECN': ecn,
                'first_price': price,
                'last_price': price,
                'first_post_hour': ts.hour + ts.minute / 60,
                'num_updates': 1,
            }
        else:
            self.today_sessions[msg_id]['last_price'] = price
            self.today_sessions[msg_id]['num_updates'] += 1

        # If new ECN_MSG_ID appeared, re-run matching for this group
        group_key = (sec_id, qty_bucket)
        if is_new_session:
            self._match_group(group_key)

        # Look up this ECN_MSG_ID's linked prior dealer
        linked_prior_id = self.dealer_links.get(msg_id)
        if linked_prior_id is None:
            # Unmatched — could be a new dealer not present yesterday
            processed = {
                'SECURITY_ID': sec_id,
                'qty_bucket': qty_bucket,
                'time_bucket': time_bucket,
                'Timestamp': ts,
                'ALT_PRICE_VALUE': price,
                'ECN_MSG_ID': msg_id,
                'ECN': ecn,
                'linked_to': None,
                'prior_price': None,
                'bps_delta': None,
                'matched': False,
            }
            self.ticks.append(processed)
            return processed

        # Get prior dealer's last price
        prior_session = next(
            (s for s in self.prior_sessions if s['ECN_MSG_ID'] == linked_prior_id
             and s['SECURITY_ID'] == sec_id and s['qty_bucket'] == qty_bucket),
            None
        )
        if prior_session is None:
            self.ticks.append({
                'SECURITY_ID': sec_id, 'qty_bucket': qty_bucket,
                'time_bucket': time_bucket, 'Timestamp': ts,
                'ALT_PRICE_VALUE': price, 'ECN_MSG_ID': msg_id,
                'ECN': ecn, 'linked_to': linked_prior_id,
                'prior_price': None, 'bps_delta': None, 'matched': False,
            })
            return None

        prior_price = prior_session['last_price']
        bps_delta = round((price - prior_price) * 100, 2)

        processed = {
            'SECURITY_ID': sec_id,
            'qty_bucket': qty_bucket,
            'time_bucket': time_bucket,
            'Timestamp': ts,
            'ALT_PRICE_VALUE': price,
            'ECN_MSG_ID': msg_id,
            'ECN': ecn,
            'linked_to': linked_prior_id,
            'prior_price': prior_price,
            'bps_delta': bps_delta,
            'matched': True,
        }
        self.ticks.append(processed)
        return processed

    def on_tick_batch(self, df):
        results = []
        for _, row in df.sort_values('Timestamp').iterrows():
            r = self.on_tick(row)
            if r is not None:
                results.append(r)
        return results

    # =================================================================
    # HUNGARIAN MATCHING (per CUSIP + qty_bucket group)
    # =================================================================
    def _match_group(self, group_key):
        """
        Run Hungarian algorithm to link today's ECN_MSG_IDs to
        yesterday's within the same (CUSIP, qty_bucket).
        """
        sec_id, qty_bucket = group_key

        # Yesterday's sessions for this group
        prev = [s for s in self.prior_sessions
                if s['SECURITY_ID'] == sec_id and s['qty_bucket'] == qty_bucket]

        # Today's sessions for this group
        curr = [s for s in self.today_sessions.values()
                if s['SECURITY_ID'] == sec_id and s['qty_bucket'] == qty_bucket]

        if not prev or not curr:
            return

        n_prev, n_curr = len(prev), len(curr)
        max_dim = max(n_prev, n_curr)

        # Build cost matrix
        HIGH_COST = 1000.0
        cost_matrix = np.full((max_dim, max_dim), HIGH_COST)

        for pi in range(n_prev):
            for ci in range(n_curr):
                cost_matrix[pi, ci] = self._match_cost(prev[pi], curr[ci])

        # Solve
        row_ind, col_ind = linear_sum_assignment(cost_matrix)

        # Apply links (clear old links for this group first)
        for s in curr:
            self.dealer_links.pop(s['ECN_MSG_ID'], None)

        for ri, ci in zip(row_ind, col_ind):
            if ri < n_prev and ci < n_curr and cost_matrix[ri, ci] < self.match_cost_threshold:
                today_id = curr[ci]['ECN_MSG_ID']
                yest_id = prev[ri]['ECN_MSG_ID']
                self.dealer_links[today_id] = yest_id

    def _match_cost(self, prev_session, curr_session):
        """
        Cost between a yesterday session and a today session.
        Lower = more likely same dealer.

        Features:
          - ECN match (same platform = strong signal)
          - Price proximity (dealer offers drift gradually)
          - Post time similarity (dealers post at similar times)
          - Update frequency similarity
        """
        cost = 0.0

        # ECN match (weight: 30%)
        ecn_penalty = 0.0 if prev_session['ECN'] == curr_session['ECN'] else 5.0
        cost += 3.0 * ecn_penalty

        # Price proximity: yesterday's last vs today's first (weight: 40%)
        price_diff_bps = abs(prev_session['last_price'] - curr_session['first_price']) * 100
        cost += 4.0 * price_diff_bps

        # Post time similarity (weight: 20%)
        time_diff = abs(prev_session['first_post_hour'] - curr_session['first_post_hour'])
        cost += 2.0 * time_diff

        # Update frequency (weight: 10%)
        update_diff = abs(prev_session['num_updates'] - curr_session['num_updates'])
        cost += 1.0 * update_diff

        return cost

    # =================================================================
    # CHART DATA (aggregated)
    # =================================================================
    def get_chart_data(self):
        """
        Aggregate linked dealers' bps deltas into time buckets per qty_bucket.
        Only matched ticks contribute.

        Returns: DataFrame with qty_bucket, time_bucket, avg_dod_bps
        """
        if not self.ticks:
            return pd.DataFrame()

        df = pd.DataFrame(self.ticks)
        df_matched = df[df['matched'] == True].copy()

        if df_matched.empty:
            return pd.DataFrame()

        # Per CUSIP per time bucket
        cusip_level = (
            df_matched.groupby(['qty_bucket', 'time_bucket', 'SECURITY_ID'])
            .agg(
                cusip_avg_bps=('bps_delta', 'mean'),
                cusip_min_bps=('bps_delta', 'min'),
                cusip_max_bps=('bps_delta', 'max'),
                cusip_ticks=('bps_delta', 'count'),
                cusip_dealers=('ECN_MSG_ID', 'nunique'),
            )
            .reset_index()
        )

        # Dealer-weighted aggregation across CUSIPs
        def weighted_agg(grp):
            w = grp['cusip_dealers']
            total_w = w.sum()
            if total_w == 0:
                w = np.ones(len(grp)) / len(grp)
            else:
                w = w / total_w
            return pd.Series({
                'avg_dod_bps': round((grp['cusip_avg_bps'] * w).sum(), 2),
                'min_dod_bps': round(grp['cusip_min_bps'].min(), 2),
                'max_dod_bps': round(grp['cusip_max_bps'].max(), 2),
                'tick_count': int(grp['cusip_ticks'].sum()),
                'num_cusips': len(grp),
                'num_dealers': int(grp['cusip_dealers'].sum()),
            })

        result = (
            cusip_level.groupby(['qty_bucket', 'time_bucket'])
            .apply(weighted_agg).reset_index()
        )
        result['_qty_sort'] = result['qty_bucket'].map(QTY_BUCKET_ORDER)
        return result.sort_values(['_qty_sort', 'time_bucket']).reset_index(drop=True)

    def get_latest_reading(self):
        cd = self.get_chart_data()
        if cd.empty:
            return {}
        latest = {}
        for qty, grp in cd.groupby('qty_bucket'):
            last = grp.sort_values('time_bucket').iloc[-1]
            latest[qty] = {
                'time_bucket': last['time_bucket'],
                'avg_dod_bps': last['avg_dod_bps'],
                'tick_count': int(last['tick_count']),
                'num_dealers': int(last['num_dealers']),
            }
        return latest

    def get_dealer_links(self):
        """Show which today ECN_MSG_ID is linked to which yesterday ECN_MSG_ID."""
        rows = []
        for today_id, yest_id in self.dealer_links.items():
            today_s = self.today_sessions.get(today_id, {})
            yest_s = next((s for s in self.prior_sessions if s['ECN_MSG_ID'] == yest_id), {})
            rows.append({
                'today_ECN_MSG_ID': today_id,
                'yesterday_ECN_MSG_ID': yest_id,
                'SECURITY_ID': today_s.get('SECURITY_ID', ''),
                'qty_bucket': today_s.get('qty_bucket', ''),
                'ECN': today_s.get('ECN', ''),
                'yesterday_last_price': yest_s.get('last_price', ''),
                'today_current_price': today_s.get('last_price', ''),
            })
        return pd.DataFrame(rows) if rows else pd.DataFrame()

    def get_unmatched_ticks(self):
        """Return ticks that couldn't be linked to a prior dealer."""
        if not self.ticks:
            return pd.DataFrame()
        df = pd.DataFrame(self.ticks)
        return df[df['matched'] == False]

    def get_raw_ticks(self):
        return pd.DataFrame(self.ticks) if self.ticks else pd.DataFrame()

    def reset_intraday(self):
        self.today_sessions = {}
        self.ticks = []
        self.dealer_links = {}
        self._matched_groups = {}

    # =================================================================
    # OUTLIER REMOVAL
    # =================================================================
    def _remove_outliers(self, df):
        clean_mask = pd.Series(True, index=df.index)
        for (sec, qty), grp in df.groupby(['SECURITY_ID', 'qty_bucket']):
            if len(grp) < 4:
                continue
            prices = grp['ALT_PRICE_VALUE']
            q1, q3 = prices.quantile(0.25), prices.quantile(0.75)
            iqr = q3 - q1
            if iqr == 0:
                continue
            lower, upper = q1 - self.outlier_iqr_mult * iqr, q3 + self.outlier_iqr_mult * iqr
            clean_mask.loc[grp[(prices < lower) | (prices > upper)].index] = False
        return df[clean_mask].reset_index(drop=True)

    # =================================================================
    # EXCEL OUTPUT
    # =================================================================
    def to_excel(self, filepath):
        cd = self.get_chart_data()
        if cd.empty:
            print("  No matched ticks to export.")
            return

        wb = Workbook()
        hdr_font = Font(bold=True, size=11, color='FFFFFF')
        hdr_fill = PatternFill('solid', fgColor='2F5496')
        thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

        # --- Sheet 1: Sentiment Data ---
        ws = wb.active
        ws.title = "Sentiment Data"
        headers = ['Qty Bucket', 'Time Bucket', 'Avg DoD (bps)', 'Min (bps)',
                   'Max (bps)', 'Ticks', 'CUSIPs', 'Dealers']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font, c.fill, c.border = hdr_font, hdr_fill, thin
            c.alignment = Alignment(horizontal='center')
        for ri, (_, r) in enumerate(cd.iterrows(), 2):
            vals = [r['qty_bucket'], r['time_bucket'], r['avg_dod_bps'],
                    r['min_dod_bps'], r['max_dod_bps'],
                    r['tick_count'], r['num_cusips'], r['num_dealers']]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=ri, column=col, value=val)
                c.border = thin
                if col in [3, 4, 5]:
                    c.number_format = '+0.00;-0.00;0.00'
                c.alignment = Alignment(horizontal='center')
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 16

        # --- Sheet 2: Chart ---
        cs = wb.create_sheet("Chart")
        qty_buckets = sorted(cd['qty_bucket'].unique(), key=lambda x: QTY_BUCKET_ORDER.get(x, 99))
        all_times = sorted(cd['time_bucket'].unique())

        cs.cell(row=1, column=1, value='Time').font = hdr_font
        cs.cell(row=1, column=1).fill = hdr_fill
        cs.cell(row=1, column=1).border = thin
        for qi, qty in enumerate(qty_buckets, 2):
            c = cs.cell(row=1, column=qi, value=qty)
            c.font, c.fill, c.border = hdr_font, hdr_fill, thin
            c.alignment = Alignment(horizontal='center')
        for ri, t in enumerate(all_times, 2):
            cs.cell(row=ri, column=1, value=t).border = thin
            for qi, qty in enumerate(qty_buckets, 2):
                match = cd[(cd['time_bucket'] == t) & (cd['qty_bucket'] == qty)]
                val = match['avg_dod_bps'].iloc[0] if not match.empty else None
                c = cs.cell(row=ri, column=qi, value=val)
                c.border = thin
                if val is not None:
                    c.number_format = '+0.00;-0.00;0.00'
        cs.column_dimensions['A'].width = 10
        for qi in range(2, len(qty_buckets) + 2):
            cs.column_dimensions[get_column_letter(qi)].width = 14

        chart = LineChart()
        chart.title = "Dealer Sentiment (Linked) — bps Δ from Prior Close"
        chart.x_axis.title = "Timestamp"
        chart.y_axis.title = "bps"
        chart.style = 10
        chart.width = 28
        chart.height = 14
        n_rows = len(all_times)
        cats = Reference(cs, min_col=1, min_row=2, max_row=1 + n_rows)
        chart.set_categories(cats)
        colors_hex = {
            '<50K': '94A3B8', '50-100K': '06B6D4', '100-150K': '8B5CF6',
            '150-200K': 'F97316', '200-500K': '10B981', '500K-1MM': 'EC4899', '1MM+': 'F59E0B',
        }
        for qi, qty in enumerate(qty_buckets, 2):
            vals = Reference(cs, min_col=qi, min_row=1, max_row=1 + n_rows)
            chart.add_data(vals, titles_from_data=True)
            chart.series[-1].graphicalProperties.line.width = 22000
            chart.series[-1].graphicalProperties.line.solidFill = colors_hex.get(qty, '64748B')
        cs.add_chart(chart, f"A{n_rows + 4}")

        # --- Sheet 3: Dealer Links ---
        dl = self.get_dealer_links()
        ws3 = wb.create_sheet("Dealer Links")
        if not dl.empty:
            link_headers = list(dl.columns)
            for col, h in enumerate(link_headers, 1):
                c = ws3.cell(row=1, column=col, value=h)
                c.font, c.fill, c.border = hdr_font, hdr_fill, thin
                c.alignment = Alignment(horizontal='center')
            for ri, (_, r) in enumerate(dl.iterrows(), 2):
                for col, h in enumerate(link_headers, 1):
                    c = ws3.cell(row=ri, column=col, value=r[h])
                    c.border = thin
                    c.alignment = Alignment(horizontal='center')
                    if 'price' in h.lower():
                        c.number_format = '0.000'
            for col in range(1, len(link_headers) + 1):
                ws3.column_dimensions[get_column_letter(col)].width = 22

        wb.save(filepath)
        print(f"  Saved: {filepath}")

    # =================================================================
    # SUMMARY
    # =================================================================
    def summary(self):
        cd = self.get_chart_data()
        if cd.empty:
            print("No matched ticks yet.")
            return

        total_ticks = len(self.ticks)
        matched_ticks = sum(1 for t in self.ticks if t['matched'])
        unmatched = total_ticks - matched_ticks

        print("=" * 70)
        print("INTRADAY SENTIMENT (Dealer-Linked)")
        print("=" * 70)
        print(f"  Dealer links:  {len(self.dealer_links)}")
        print(f"  Ticks:         {matched_ticks} matched / {unmatched} unmatched / {total_ticks} total")
        avail = sorted(cd['qty_bucket'].unique(), key=lambda x: QTY_BUCKET_ORDER.get(x, 99))
        print(f"  Qty buckets:   {avail}")

        print(f"\n  {'Qty':<12} {'Open':>7} {'Latest':>7} {'Low':>7} {'High':>7} {'Ticks':>6}")
        print(f"  {'─'*12} {'─'*7} {'─'*7} {'─'*7} {'─'*7} {'─'*6}")
        for _, grp in cd.groupby('_qty_sort'):
            g = grp.sort_values('time_bucket')
            print(f"  {g.iloc[0]['qty_bucket']:<12} {g.iloc[0]['avg_dod_bps']:>+7.2f} "
                  f"{g.iloc[-1]['avg_dod_bps']:>+7.2f} {g['min_dod_bps'].min():>+7.2f} "
                  f"{g['max_dod_bps'].max():>+7.2f} {int(g['tick_count'].sum()):>6}")


# =============================================================================
# EXAMPLE USAGE
# =============================================================================
if __name__ == '__main__':

    INPUT_PATH = r"C:\Users\savya\OneDrive\Desktop\loading\test_dealer_extended_3secids.xlsx"
    OUTPUT_PATH = r"C:\Users\savya\OneDrive\Desktop\loading\dealer_sentiment_linked.xlsx"

    df = pd.read_excel(INPUT_PATH)
    df['ECN_MSG_ID'] = df['ECN_MSG_ID'].astype(str)
    df['Date'] = df['Timestamp'].dt.date

    dates = sorted(df['Date'].unique())
    prior_day = dates[-2]
    today = dates[-1]

    print(f"Simulating: prior day = {prior_day}, today = {today}\n")

    # 1. PRE-MARKET
    model = IntraDaySentiment(bucket_minutes=5)
    model.pre_market(df[df['Date'] == prior_day])

    # 2. FEED TODAY'S TICKS
    todays_ticks = df[df['Date'] == today].sort_values('Timestamp')
    print(f"\nProcessing {len(todays_ticks)} ticks...\n")

    for _, tick in todays_ticks.iterrows():
        result = model.on_tick(tick)
        if result and result['matched']:
            print(f"  {result['time_bucket']}  {result['SECURITY_ID']}  "
                  f"{result['qty_bucket']:>10}  {result['bps_delta']:>+6.2f} bps  "
                  f"(linked: {result['ECN_MSG_ID']} -> {result['linked_to']})")
        elif result:
            print(f"  {result['time_bucket']}  {result['SECURITY_ID']}  "
                  f"{result['qty_bucket']:>10}  UNMATCHED  ({result['ECN_MSG_ID']})")

    # 3. RESULTS
    print()
    model.summary()

    print("\nDealer links:")
    dl = model.get_dealer_links()
    if not dl.empty:
        print(dl.to_string(index=False))

    print("\nLatest reading:")
    for qty, r in model.get_latest_reading().items():
        print(f"  {qty:>10}: {r['avg_dod_bps']:>+.2f} bps at {r['time_bucket']}")

    print("\nWriting Excel...")
    model.to_excel(OUTPUT_PATH)

    # END OF DAY
    # model.reset_intraday()
    # model.pre_market(df[df['Date'] == today])
