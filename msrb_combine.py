# msrb_combine.py
# Streamlit app that:
# 1) Spawns a writer thread that appends live MSRB yields to Excel
# 2) Spawns a reader thread that keeps an in-memory pandas DataFrame updated
# 3) Plots the live DataFrame with Plotly and auto-refreshes the UI

import os
import time
import threading
from datetime import datetime
import random

import pandas as pd
from openpyxl import Workbook, load_workbook
import streamlit as st
import plotly.express as px

# =========================
# ---- CONFIG DEFAULTS ----
# =========================
FILE_PATH_DEFAULT = r"C:\Users\savya\OneDrive\MSRB_Yield.xlsx"
SHEET_NAME = "Sheet1"
DEFAULT_WRITE_INTERVAL = 30  # seconds
DEFAULT_READ_INTERVAL = 5    # seconds
DEFAULT_REFRESH_MS = 3000    # Streamlit UI refresh interval
DEFAULT_SHOW_LAST_ROWS = 500 # rows in chart (0 = all)
MEAN_YIELD = 3.0
SIGMA_YIELD = 0.05

# Shared state (protected by a lock)
latest_df_lock = threading.Lock()
latest_df = pd.DataFrame(columns=["Timestamp", "MSRB_Yield"])

# Graceful stop signal if needed later
stop_event = threading.Event()

# =========================
# ---- THREAD TARGETS  ----
# =========================
def writer_task(file_path: str, sheet_name: str, interval_sec: int):
    """Append a new (Timestamp, MSRB_Yield) tick to Excel every interval_sec."""
    # Ensure file exists with headers
    if not os.path.exists(file_path):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            ws.append(["Timestamp", "MSRB_Yield"])
            wb.save(file_path)
            wb.close()
            print("[Writer] Created workbook with headers.")
        except Exception as e:
            print("[Writer] Failed to create workbook:", e)

    print(f"[Writer] Running. Interval: {interval_sec}s -> {file_path}")
    while not stop_event.is_set():
        ts = datetime.now()
        yld = random.gauss(MEAN_YIELD, SIGMA_YIELD)
        wrote = False

        # Retry a few times in case OneDrive/Excel locks the file
        for _ in range(5):
            try:
                wb = load_workbook(file_path)
                ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
                ws.append([ts, round(yld, 6)])
                wb.save(file_path)
                wb.close()
                wrote = True
                break
            except PermissionError:
                time.sleep(0.7)
            except Exception as e:
                print("[Writer] Error:", e)
                time.sleep(0.7)

        if wrote:
            print(f"[Writer] Appended: {ts} | Yield: {yld:.4f}")
        else:
            print("[Writer] Could not write this tick (locked). Will try next interval.")

        time.sleep(interval_sec)

def reader_task(file_path: str, sheet_name: str, interval_sec: int):
    """Continuously read Excel into latest_df (in memory) every interval_sec."""
    global latest_df
    print(f"[Reader] Running. Interval: {interval_sec}s <- {file_path}")
    while not stop_event.is_set():
        try:
            # Small retry loop for transient locks
            attempts = 5
            df_local = None
            for i in range(attempts):
                try:
                    df_local = pd.read_excel(file_path, sheet_name=sheet_name)
                    break
                except PermissionError:
                    time.sleep(0.5)
                except Exception as e:
                    if i == attempts - 1:
                        raise e
                    time.sleep(0.5)

            if df_local is not None and not df_local.empty:
                if "Timestamp" in df_local.columns:
                    df_local["Timestamp"] = pd.to_datetime(df_local["Timestamp"], errors="coerce")
                    df_local = df_local.dropna(subset=["Timestamp"])

                with latest_df_lock:
                    latest_df = df_local

                print(f"[Reader] Loaded {len(df_local)} rows @ {datetime.now():%H:%M:%S}")
        except Exception as e:
            print("[Reader] Error reading Excel:", e)

        time.sleep(interval_sec)

# =========================
# ------- STREAMLIT -------
# =========================
st.set_page_config(page_title="Live MSRB Yield (Writer + Reader + Plot)", layout="wide")
st.title("📈 Live MSRB Yield — Combined Writer/Reader Dashboard")
st.caption("This app writes live ticks to Excel and simultaneously reads them into a shared DataFrame for plotting.")

with st.sidebar:
    st.subheader("Settings")
    file_path = st.text_input("Excel path", value=FILE_PATH_DEFAULT)
    write_interval = st.number_input("Writer interval (sec)", min_value=1, max_value=120, value=DEFAULT_WRITE_INTERVAL, step=1)
    read_interval = st.number_input("Reader interval (sec)", min_value=1, max_value=60, value=DEFAULT_READ_INTERVAL, step=1)
    refresh_ms = st.number_input("UI refresh (ms)", min_value=500, max_value=10000, value=DEFAULT_REFRESH_MS, step=500)
    show_last_n = st.number_input("Show last N rows (0 = all)", min_value=0, max_value=500000, value=DEFAULT_SHOW_LAST_ROWS, step=100)
    st.write("Tip: keep the Excel file closed while streaming to avoid locks.")

# Start threads once per session
if "threads_started" not in st.session_state:
    st.session_state.threads_started = False

if not st.session_state.threads_started:
    # Validate path directory exists
    dir_ok = os.path.isdir(os.path.dirname(file_path)) if os.path.dirname(file_path) else True
    if not dir_ok:
        st.error(f"Directory does not exist: {os.path.dirname(file_path)}")
        st.stop()

    # Launch writer & reader threads
    t_writer = threading.Thread(
        target=writer_task, args=(file_path, SHEET_NAME, write_interval), daemon=True, name="Writer-T1"
    )
    t_reader = threading.Thread(
        target=reader_task, args=(file_path, SHEET_NAME, read_interval), daemon=True, name="Reader-T2"
    )
    t_writer.start()
    t_reader.start()

    st.session_state.threads_started = True
    st.success("✅ Writer and Reader threads started.")

# Auto-refresh the Streamlit UI
st.autorefresh(interval=refresh_ms, key="msrb_autorefresh")

# Read the in-memory DataFrame and plot
with latest_df_lock:
    df_plot = latest_df.copy()

if df_plot.empty:
    st.info("Waiting for data… (ensure writer thread can create/append to the Excel file)")
else:
    if show_last_n and show_last_n > 0 and len(df_plot) > show_last_n:
        df_plot = df_plot.tail(show_last_n)

    fig = px.line(
        df_plot,
        x="Timestamp",
        y="MSRB_Yield",
        title="Live MSRB Yields",
        markers=True,
    )
    fig.update_layout(
        xaxis_title="Time",
        yaxis_title="Yield (%)",
        margin=dict(l=10, r=10, t=40, b=10),
    )

    st.plotly_chart(fig, use_container_width=True)
    st.caption(f"Rows shown: {len(df_plot)} (Total live rows: {len(latest_df)})")

    st.write("Latest 10 ticks")
    st.dataframe(latest_df.tail(10), use_container_width=True)

st.write(f"⏱️ Last UI refresh: {pd.Timestamp.now():%Y-%m-%d %H:%M:%S}")
