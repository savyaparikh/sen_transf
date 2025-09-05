import time
import random
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

def live_msrbyield_to_excel(file_path=r"C:\Users\savya\OneDrive\MSRB_Yield.xlsx",
                            mean=3.0, sigma=0.05, interval_sec=30, sheet_name="Sheet1"):
    """
    Append new MSRB yield ticks to Excel every 30 sec
    without deleting existing data.
    """

    # If file doesn’t exist, create it with headers
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(["Timestamp", "MSRB_Yield"])
        wb.save(file_path)
        wb.close()
        print("Created new Excel file with headers.")

    print(f"Appending new ticks to {file_path} every {interval_sec} seconds. Ctrl+C to stop.")

    while True:
        ts = datetime.now()
        yld = random.gauss(mean, sigma)

        try:
            wb = load_workbook(file_path)
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
            ws.append([ts, round(yld, 6)])  # add at bottom
            wb.save(file_path)
            wb.close()
            print(f"Appended: {ts} | Yield: {yld:.4f}")
        except Exception as e:
            print("⚠️ Error writing file (maybe locked by Excel):", e)

        time.sleep(interval_sec)

# Run it:
live_msrbyield_to_excel()
